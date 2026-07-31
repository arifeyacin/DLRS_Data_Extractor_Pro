"""
OpenPyXL Excel Exporter Module for DLRS Data Extractor Pro
Generates styled Excel spreadsheets with headers, auto-column widths, and grid formatting.
"""

import os
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.logger import get_system_logger, get_error_logger

class ExcelExporter:
    """Generates styled Excel workbooks from land records."""

    def __init__(self, output_dir: str = "./Output/Excel"):
        self.output_dir = os.path.abspath(output_dir)
        os.makedirs(self.output_dir, exist_ok=True)
        self.logger = get_system_logger()
        self.error_logger = get_error_logger()

    def export(self, records: List[Dict[str, Any]], filename: str = "dlrs_land_records.xlsx") -> str:
        """Export records list to styled Excel workbook."""
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        target_path = os.path.join(self.output_dir, filename)
        self.logger.info(f"Exporting {len(records)} records to Excel spreadsheet: {target_path}")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Vested Property Records"
            ws.views.sheetView[0].showGridLines = True

            if not records:
                wb.save(target_path)
                return target_path

            # Field Headers
            headers = [
                "Record ID", "Gazette No", "District (জেলা)", "Upazila (উপজেলা)",
                "Mouza (মৌজা)", "Khatian (খতিয়ান)", "Dag (দাগ)", "Owner Name (মালিকের নাম)",
                "Pub Date", "Area", "Land Type", "Remarks", "Source PDF"
            ]
            field_keys = [
                "record_id", "gazette_number", "district", "upazila", "mouza",
                "khatian", "dag", "owner_name", "publication_date", "area",
                "land_type", "remarks", "source_pdf"
            ]

            # Header Styling
            header_fill = PatternFill(start_color="00A63E", end_color="00A63E", fill_type="solid")
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0')
            )

            # Write Header Row
            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = thin_border

            ws.row_dimensions[1].height = 28

            # Write Data Rows
            data_font = Font(name="Segoe UI", size=10)
            data_align = Alignment(vertical="center")

            for r_idx, record in enumerate(records, start=2):
                row_data = [str(record.get(k, "")) for k in field_keys]
                ws.append(row_data)

                for c_idx in range(1, len(row_data) + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = thin_border

            # Auto-fit Column Widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

            # Freeze Header Row
            ws.freeze_panes = "A2"

            wb.save(target_path)
            self.logger.info(f"Excel export completed successfully: {target_path}")
            return target_path
        except Exception as e:
            self.error_logger.error(f"Failed to export Excel file {target_path}: {e}")
            return ""
